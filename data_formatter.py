import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import dateutil.parser
from datetime import datetime

def parse_date(date_string):
    try:
        if 'noon' in date_string.lower():
            date_part = date_string.lower().replace('noon', '').strip()
            parsed_date = dateutil.parser.parse(date_part)
            return parsed_date.replace(hour=12, minute=0)
        elif 'midnight' in date_string.lower():
            date_part = date_string.lower().replace('midnight', '').strip()
            parsed_date = dateutil.parser.parse(date_part)
            return parsed_date.replace(hour=0, minute=0)
        else:
            return dateutil.parser.parse(date_string)
    except Exception as e:
        print(f"Failed to parse date: {date_string}. Error: {str(e)}")
        return pd.NaT

def format_date(date):
    if pd.isnull(date):
        return ''
    try:
        return date.strftime('%Y-%m-%d %H:%M')
    except Exception as e:
        print(f"Failed to format date: {date}. Error: {str(e)}")
        return ''

def apply_table_styles(ws, start_cell, end_cell, style_name="TableStyleMedium2"):
    tab = Table(displayName=f"{ws.title.replace(' ', '')}Table", ref=f"{start_cell}:{end_cell}")
    style = TableStyleInfo(name=style_name, showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def adjust_column_widths(ws):
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = max_length + 2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

def create_header_style():
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    header_style.fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    return header_style

def create_data_style():
    data_style = NamedStyle(name="data_style")
    data_style.font = Font(name='Calibri', size=11)
    data_style.alignment = Alignment(horizontal='left', vertical='center')
    return data_style

# Read and process data
df = pd.read_csv('space_launches.csv')
df['date'] = df['date'].apply(parse_date)
df = df.sort_values('date')
df['formatted_date'] = df['date'].apply(format_date)
df['year'] = df['date'].dt.year
df['month'] = df['date'].dt.to_period('M')
df['success'] = df['status'].str.contains('Successful', case=False, na=False).astype(int)

# Create workbook and styles
wb = Workbook()
ws = wb.active
ws.title = "Space Launches"

header_style = create_header_style()
data_style = create_data_style()
wb.add_named_style(header_style)
wb.add_named_style(data_style)

# Main Space Launches sheet
columns_to_write = ['name', 'formatted_date', 'status', 'provider', 'rocket', 'mission', 'pad']
headers = ['Name', 'Date', 'Status', 'Provider', 'Rocket', 'Mission', 'Launch Pad']

ws.append(headers)
for cell in ws[1]:
    cell.style = header_style

for r_idx, row in enumerate(dataframe_to_rows(df[columns_to_write], index=False, header=False), start=2):
    ws.append(row)
    for cell in ws[r_idx]:
        cell.style = data_style

apply_table_styles(ws, "A1", f"G{ws.max_row}")
adjust_column_widths(ws)
ws.freeze_panes = "A2"

# Yearly Launch Statistics
yearly_stats = df.groupby('year').agg({
    'success': ['sum', 'count'],
    'provider': 'nunique',
    'rocket': 'nunique'
}).reset_index()
yearly_stats.columns = ['Year', 'Successful Launches', 'Total Launches', 'Unique Providers', 'Unique Rockets']
yearly_stats['Success Rate'] = yearly_stats['Successful Launches'] / yearly_stats['Total Launches']

stats_sheet = wb.create_sheet(title="Yearly Statistics")
stats_sheet.append(['Year', 'Successful Launches', 'Total Launches', 'Success Rate', 'Unique Providers', 'Unique Rockets'])
for cell in stats_sheet[1]:
    cell.style = header_style

for _, row in yearly_stats.iterrows():
    stats_sheet.append(row.tolist())
    for cell in stats_sheet[stats_sheet.max_row]:
        cell.style = data_style

apply_table_styles(stats_sheet, "A1", f"F{stats_sheet.max_row}")
adjust_column_widths(stats_sheet)

# Launch Trend Analysis
trend_sheet = wb.create_sheet(title="Launch Trends")
monthly_launches = df.groupby('month').size().reset_index(name='Launches')
monthly_launches['month'] = monthly_launches['month'].astype(str)

trend_sheet.append(['Month', 'Launches'])
for cell in trend_sheet[1]:
    cell.style = header_style

for _, row in monthly_launches.iterrows():
    trend_sheet.append(row.tolist())
    for cell in trend_sheet[trend_sheet.max_row]:
        cell.style = data_style

apply_table_styles(trend_sheet, "A1", f"B{trend_sheet.max_row}")
adjust_column_widths(trend_sheet)

line_chart = LineChart()
line_chart.title = "Monthly Launch Trend"
line_chart.x_axis.title = "Month"
line_chart.y_axis.title = "Number of Launches"
data = Reference(trend_sheet, min_col=2, min_row=1, max_row=len(monthly_launches)+1)
cats = Reference(trend_sheet, min_col=1, min_row=2, max_row=len(monthly_launches)+1)
line_chart.add_data(data, titles_from_data=True)
line_chart.set_categories(cats)
trend_sheet.add_chart(line_chart, "D2")
line_chart.height = 15  # Adjust as needed
line_chart.width = 30   # Adjust as needed

# Provider Performance Analysis
provider_sheet = wb.create_sheet(title="Provider Performance")
provider_stats = df.groupby('provider').agg({
    'success': ['sum', 'count'],
    'rocket': 'nunique'
}).reset_index()
provider_stats.columns = ['Provider', 'Successful Launches', 'Total Launches', 'Unique Rockets']
provider_stats['Success Rate'] = provider_stats['Successful Launches'] / provider_stats['Total Launches']
provider_stats = provider_stats.sort_values('Total Launches', ascending=False).head(10)

provider_sheet.append(['Provider', 'Successful Launches', 'Total Launches', 'Success Rate', 'Unique Rockets'])
for cell in provider_sheet[1]:
    cell.style = header_style

for _, row in provider_stats.iterrows():
    provider_sheet.append(row.tolist())
    for cell in provider_sheet[provider_sheet.max_row]:
        cell.style = data_style

apply_table_styles(provider_sheet, "A1", f"E{provider_sheet.max_row}")
adjust_column_widths(provider_sheet)

bar_chart = BarChart()
bar_chart.title = "Top 10 Providers by Total Launches"
bar_chart.x_axis.title = "Provider"
bar_chart.y_axis.title = "Number of Launches"
data = Reference(provider_sheet, min_col=2, max_col=3, min_row=1, max_row=11)
cats = Reference(provider_sheet, min_col=1, min_row=2, max_row=11)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
provider_sheet.add_chart(bar_chart, "G2")
bar_chart.height = 15  # Adjust as needed
bar_chart.width = 30   # Adjust as needed

# Launch Site Analysis
site_sheet = wb.create_sheet(title="Launch Sites")
site_stats = df.groupby('pad').agg({
    'success': ['sum', 'count'],
    'provider': 'nunique'
}).reset_index()
site_stats.columns = ['Launch Site', 'Successful Launches', 'Total Launches', 'Unique Providers']
site_stats['Success Rate'] = site_stats['Successful Launches'] / site_stats['Total Launches']
site_stats = site_stats.sort_values('Total Launches', ascending=False).head(10)

site_sheet.append(['Launch Site', 'Successful Launches', 'Total Launches', 'Success Rate', 'Unique Providers'])
for cell in site_sheet[1]:
    cell.style = header_style

for _, row in site_stats.iterrows():
    site_sheet.append(row.tolist())
    for cell in site_sheet[site_sheet.max_row]:
        cell.style = data_style

apply_table_styles(site_sheet, "A1", f"E{site_sheet.max_row}")
adjust_column_widths(site_sheet)

pie_chart = PieChart()
pie_chart.title = "Top 10 Launch Sites by Total Launches"
data = Reference(site_sheet, min_col=3, min_row=1, max_row=11)
labels = Reference(site_sheet, min_col=1, min_row=2, max_row=11)
pie_chart.add_data(data, titles_from_data=True)
pie_chart.set_categories(labels)
site_sheet.add_chart(pie_chart, "G2")
pie_chart.height = 15  
pie_chart.width = 30   

# Save workbook
wb.save("space_launches_advanced_report.xlsx")
print("Excel file 'space_launches_advanced_report.xlsx' has been generated.")

# Print statistics
print("\nAnalysis Summary:")
print(f"Total launches analyzed: {len(df)}")
print(f"Date range: {df['date'].min().year} to {df['date'].max().year}")
print(f"Overall success rate: {df['success'].mean():.2%}")
print(f"Total unique providers: {df['provider'].nunique()}")
print(f"Total unique rockets: {df['rocket'].nunique()}")
print(f"Total launch sites: {df['pad'].nunique()}")
